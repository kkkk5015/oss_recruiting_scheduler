import webbrowser
import openpyxl as xl
import datetime as d
import requests
import datetime
from bs4 import BeautifulSoup
import urllib
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
'''
    기능 저장, 분류, ...
    
    첫화면 기능 선택 
        저장 시 필요한 요소 
            회사명, 마감일, 코테 유무(날짜), 회사지원url, 코테 지원 언어,...
            
        분류 
            지원 날짜별,  
'''
def create(): # 엑셀파일 만들어서 구분하기

    try :
        xl.load_workbook('Company Schedules.xlsx')
    except FileNotFoundError as f :
        wb = Workbook()
        ws = wb.active
        # 제목
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        ws.merge_cells('A1:E1')
        ws['A1'] = 'Company Schedules' # 변경하셔도 됩니다 (한국어로)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws['A1'].font = Font(bold=True, color='FF000000', size=15) # color is aRGB
        ws.cell(row=1, column=5).border = thin_border
        # 구분
        blackFill = PatternFill(fill_type='solid', start_color='FF000000', end_color='FF000000')
        ws['A2'] = "회사명"
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(bold=True, color='FFFFFFFF', size=13)
        ws['A2'].fill = blackFill
        ws['B2'] = "마감일"
        ws['B2'].alignment = Alignment(horizontal='center')
        ws['B2'].font = Font(bold=True, color='FFFFFFFF', size=13)
        ws['B2'].fill = blackFill
        ws['C2'] = "코테 유무(날짜)"
        ws['C2'].font = Font(bold=True, color='FFFFFFFF', size=13)
        ws['C2'].fill = blackFill
        ws['D2'] = '회사지원url'
        ws['D2'].alignment = Alignment(horizontal='center')
        ws['D2'].font = Font(bold=True, color='FFFFFFFF', size=13)
        ws['D2'].fill = blackFill
        ws['E2'] = '코테 지원 언어'
        ws['E2'].font = Font(bold=True, color='FFFFFFFF', size=13)
        ws['E2'].fill = blackFill

        ws['G2'] = "Last Update"
        ws['G2'].alignment = Alignment(horizontal='center')
        ws['G2'].font = Font(bold=True, color='FFFFFFFF', size=13)
        ws['G2'].fill = blackFill
        # ws.cell(row=3, column=7).border = thin_border
        ws.column_dimensions['C'].width = 17
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['G'].width = 20

        # ht = input("엑셀파일 만들기 위해 저장할 위치 입력해주세요 : ")
        wb.save('Company Schedules.xlsx')
        print()
        print("현재 파일 위치에 'Company Schedules'라는 엑셀 파일을 만들었습니다")
        return

    if xl.load_workbook('Company Schedules.xlsx') :
        print()
        an = input("이미 생성된 파일이 존재합니다. 기존 액셀에 이어서 하시겠습니까? (y/n) : ")
        if an == 'y' :
            xlsave()


def xlsave() : # 액셀 저장 파트

    # #print("일정을 저장할 엑셀파일의 주소를 적어주세요 : ",sep='',end='')
    # ht = input("일정을 저장할 엑셀파일의 주소를 적어주세요 : ")
    try :
        xl.load_workbook('Company Schedules.xlsx')
    except FileNotFoundError as f:
        print()
        print('일정 저장용 파일이 존재하지 않습니다.')
        create()

    while True :
        print()
        ent = []
        company_name = str(input("지원 하실 회사의 이름을 적어주세요(첫 화면으로 돌아가려면 *을 적어주제요) : "))
        if company_name == '*' :
            break
        d_day = str(input("지원 마감일을 적어주세요(예시 : 09/24) : "))
        code_day = str(input("코딩테스트 날짜를 적어주세요(예시: 09/27, 코딩 테스트가 없다면 x입력) : "))
        com_url = str(input("회사 지원 홈페이지 주소를 적어주세요 : "))
        code_lan = str(input("코딩테스트 언어를 모두 적어 주세요, 테스트가 없다면 x입력 (예시 : java, c++, python) : "))
        ent = [company_name,d_day,code_day,com_url,code_lan]
        wb = xl.load_workbook('Company Schedules.xlsx')
        ws = wb['Sheet']
        # 혹시 필요하다면 sheet 값도 받아야함
        ws.append(ent)
        ws['G3'] = datetime.datetime.today()
        wb.save('Company Schedules.xlsx')
        print()
        print("지원 회사가 저장되었습니다. ")
        print()



def xlout() :  # 조회 기능 파트
    # 기능 지원 회사 나열, ㅇ
    print()
    print("1. 뒤로가기")
    print("2. 회사 정보 조회하기")
    print("3. 회사 지원홈페이지 가기")
    print("4. 지원 마감된 제외하고 회사 나열하기")
    print("5. 코딩테스트가 있는 지원회사 나열하기")
    print("")
    a = input("원하시는 업무 번호를 적어주세요(나가기는 * 입력) : ")

    
    if a == '1' :
        return
    elif a == '2' :
        xlcominfo()
    elif a == '3' :
        xlrecurl()
    elif a == '4' :
        xlfinished()
    elif a == '5' :
        xlcodetest()
        


def xlrecurl() :
    print()
    # ht = input("저장된 엑셀파일의 주소를 적어주세요 : ")
    wb = xl.load_workbook('Company Schedules.xlsx')
    # 필요하다면 sheet 값도 받아야한다.
    ws = wb['Sheet']
    col_value = []
    for col in ws.columns :
        col_l = []
        for cell in col :
            col_l.append(cell.value)
        col_value.append(col_l)
    print()
    for i, com in enumerate(col_value[0][2:]) :
        print('{}. {}'.format(i,com))
    print()
    comnum = int(input("원하는 회사의 번호를 입력해주세요 : "))
    print()
    comurl = col_value[3][comnum]
    webbrowser.open(comurl)


def xlcominfo() : #회사 정보 가져오기
    print()
    comname = input("조회할 기업명을 입력해주세요 : ")
    cominfo = []
    temp = []
    #comname = '파수닷컴'
    cominfourl = 'http://www.jobkorea.co.kr/Salary/Index?coKeyword=' + comname + '&tabindex=0&indsCtgrCode=&indsCode=&jobTypeCode=&haveAGI=0&orderCode=2&coPage=1#salarySearchCompany'
    # mykey = {'keyword' : comname }
    r = requests.get(cominfourl)
    soup = BeautifulSoup(r.text, 'html.parser')
    keyword = soup.select('.salaryCompanyList .container .list li a')
    #print(keyword)
    for word in keyword:
        w = list(map(str,word.get_text().split('\n')))
        temp.append(w)

        #temp = list(map(str, word.get_text().split('\n')))
    #print(temp)
    for c in temp :
        t = []
        for k in c :
            if k in t or k == '' :
                continue
            elif k == '좋아요' or k == '채용중':
                continue
            t.append(k)

        cominfo.append(t)

    for com in cominfo :
        for p, line in enumerate(com) :
            print(line)
            if p == 0 :
                print()
        print()
        print('-------------------------------')
        print()

        


    '''
    print("기업명 : "+cominfo[0])
    print("회사직무 : "+cominfo[3])
    print(cominfo[4])
    print(cominfo[5])
    print("평균연봉 : "+cominfo[6])
    '''


def xlfinished() : #마감된 회사 출력
    try :
        xl.load_workbook('Company Schedules.xlsx')
    except FileNotFoundError as f:
        print()
        print('일정 저장용 파일이 존재하지 않습니다.')
        create()

    wb = xl.load_workbook('Company Schedules.xlsx')
    ws = wb['Sheet']

    print()
    i = 2 #카테고리 행 스킵
    for r in ws.rows :
        if(i > 0) :
            i -= 1
            continue
        
        mon = str(r[1].value).split('/')[0]
        day = str(r[1].value).split('/')[1]
        now = datetime.datetime.now()
        deadline = datetime.datetime.strptime(str(now.year)+mon+day+'235959', '%Y%m%d%H%M%S')
            
        if(datetime.datetime.now() <= deadline) :
            print('회사명:'+str(r[0].value)+'\t마감 날짜:'+r[1].value)


def xlcodetest() : #코테가 있는 회사 출력
    try :
        xl.load_workbook('Company Schedules.xlsx')
    except FileNotFoundError as f:
        print()
        print('일정 저장용 파일이 존재하지 않습니다.')
        create()

    wb = xl.load_workbook('Company Schedules.xlsx')
    ws = wb['Sheet']

    print()
    i = 2
    for r in ws.rows :
        if(i > 0) :
            i -= 1
            continue

        if(str(r[2].value) != 'x' and str(r[2].value) != 'X') : #코테가 있으면서 마감이 남은 회사 출력
            mon = str(r[1].value).split('/')[0]
            day = str(r[1].value).split('/')[1]
            now = datetime.datetime.now()
            deadline = datetime.datetime.strptime(str(now.year)+mon+day+'235959', '%Y%m%d%H%M%S')
            
            if(datetime.datetime.now() <= deadline) :
                print('회사명:'+str(r[0].value)+'\t마감 날짜:'+str(r[1].value)+'    코테 날짜:'+r[2].value+'    언어:'+r[4].value)


while True :
    print()
    print("1. 일정 저장 액셀파일 만들기")
    print("2. 지원 일정 저장하기")
    print("3. 지원 일정 조회하기")
    print("원하시는 업무 번호를 적어주세요(나가기는 * 입력) : ", sep="",end="")
    st = input()
    if st == '*' :
        break
    if st == '1' :
        create()
    if st == '2' :
        xlsave()
    if st == '3' :
        xlout()







